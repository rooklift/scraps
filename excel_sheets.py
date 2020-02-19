# Totally untested code. The point was to turn a bunch of
# tables in 1 sheet into multiple sheets, with an index
# sheet as well (with internal links).

import copy, sys

print("Loading openpyxl, this takes ages...")
import openpyxl
print("Finished loading openpyxl.")

class NotHeader(Exception):
	pass

def main():

	if len(sys.argv) < 2:
		print("Need filename.")
		return

	infile = sys.argv[-1]
	inbook = openpyxl.load_workbook(infile)

	sheet_names = inbook.get_sheet_names()

	if len(sheet_names) != 1:
		print("Expected 1 sheet, found {}.".format(len(sheet_names)))
		return

	source = inbook.get_sheet_by_name(sheet_names[0])

	header_rows = []

	max_column = source.max_column
	max_row = source.max_row

	for y in range(1, max_row + 1):

		try:
			possible_header = source.cell(column = 1, row = y)

			if possible_header.value is None:
				raise NotHeader

			for x in range(2, max_column + 1):
				cell = source.cell(column = x, row = y)
				if cell.value is not None:
					raise NotHeader

			header_rows.append(y)

		except NotHeader:
			continue

	if len(header_rows) < 1:
		print("Found no headers.")
		return

	outbook = openpyxl.Workbook()
	link_sheet = outbook.active			# The first, pre-existing sheet in the new book

	ranges = []

	for n in range(len(header_rows) - 1):
		ranges.append([header_rows[n], header_rows[n + 1]])

	ranges.append([header_rows[-1], max_row + 1])

	count = 0

	for r in ranges:

		count += 1
		target = outbook.create_sheet()
		header_text = source.cell(column = 1, row = r[0]).value
		link_cell = link_sheet.cell(column = 1, row = count)
		link_cell.value = '=HYPERLINK("#{}!A1", "{}")'.format(target.title, header_text)

		for y in range(r[0], r[1]):

			for x in range(1, max_column + 1):

				sc = source.cell(column = x, row = y)
				tc = target.cell(column = x, row = y + 1 - r[0])

				tc.value = sc.value
				tc.font = copy.copy(sc.font)

	outbook.save("{}.output.xlsx".format(infile))


main()
