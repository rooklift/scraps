from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import os

soft_green = PatternFill(start_color='DBF2C8', end_color='DBF2C8', fill_type='solid')
yellow = PatternFill(start_color='FFFFD7', end_color='FFFFD7', fill_type='solid')
red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
blank = PatternFill(fill_type=None)

def get_available_filename(base_path):
    if not os.path.exists(base_path):
        return base_path

    name, ext = os.path.splitext(base_path)
    counter = 1

    while True:
        new_path = f"{name} ({counter}){ext}"
        if not os.path.exists(new_path):
            return new_path
        counter += 1

def compare_sheets(file1_path, file2_path):
    wb1 = load_workbook(file1_path)
    wb2 = load_workbook(file2_path)

    ws1 = wb1.active
    ws2 = wb2.active

    for row in range(5, ws1.max_row + 1):
        for col in range(4, ws1.max_column + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)

            if cell1.value is None and cell2.value is None:
                cell1.fill = blank
            elif cell1.value == cell2.value:
                cell1.fill = soft_green
            elif cell1.value is not None and cell2.value is None:
                cell1.fill = yellow
            elif cell1.value is None and cell2.value is not None:
                cell1.fill = yellow
            elif cell1.value != cell2.value:
                cell1.fill = red

    basename1 = os.path.basename(file1_path).removesuffix('.xlsx')
    basename2 = os.path.basename(file2_path).removesuffix('.xlsx')
    base_output_path = f"Comparison - {basename1} - {basename2}.xlsx"

    output_path = get_available_filename(base_output_path)
    wb1.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py file1.xlsx file2.xlsx")
        sys.exit(1)

    compare_sheets(sys.argv[1], sys.argv[2])